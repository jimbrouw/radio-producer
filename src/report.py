import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def generate_excel_report(data, output_path):
    """
    Generates a formatted Excel report from the provided data.
    data: List of dictionaries representing identified tracks/segments.
    output_path: Path to save the .xlsx file.
    """
    if not data:
        logging.warning("No data to report.")
        return

    df = pd.DataFrame(data)
    
    # Ensure all columns exist even if empty
    required_columns = [
        'Filename', 'Timestamp', 'Title', 'Artist', 'Album', 
        'Label', 'Year', 'ISRC', 'Confidence'
    ]
    for col in required_columns:
        if col not in df.columns:
            df[col] = ''
            
    # Reorder columns
    df = df[required_columns]
    
    try:
        # standardpandas write
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Apply formatting
        _apply_conditional_formatting(output_path)
        logging.info(f"Report generated successfully: {output_path}")
        
    except Exception as e:
        logging.error(f"Failed to generate report: {e}")

def _apply_conditional_formatting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Iterate over rows (min_row=2 to skip header)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=9):
        # Extract cell values by column index (1-based)
        # 3=Title, 4=Artist, 6=Label, 8=ISRC
        title_cell = row[2]  # C
        label_cell = row[5]  # F
        isrc_cell = row[7]   # H
        
        fill_to_apply = None
        
        has_title = bool(title_cell.value)
        has_label = bool(label_cell.value)
        has_isrc = bool(isrc_cell.value)
        
        if not has_title:
             # Red: No track identified (assuming we log "Unknown" rows, otherwise this might just be empty)
             # If we only log found tracks, this case might be rare unless we force logic for "Unknown"
             # For now, let's say if Title is empty/missing, it's Red.
             fill_to_apply = red_fill
        elif has_title and has_label and has_isrc:
            # Green: High confidence, all data present
            fill_to_apply = green_fill
        else:
            # Yellow: Track found, but missing metadata
            fill_to_apply = yellow_fill
            
        if fill_to_apply:
            for cell in row:
                cell.fill = fill_to_apply

    wb.save(file_path)
